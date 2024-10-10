import os
import sys
import re

import pyexcel as p
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# folder_path = os.getcwd() #用此方式获取当前工作目录在打包exe后，莫名变成系统用户的主目录，并没有获取当前目录
# folder_path = os.path.dirname(os.path.abspath(sys.argv[0]))
# folder_path = "D:\\liaojq\\wechat\\20240606考勤"
# folder_path = "C:\\Users\\Administrator\\Desktop\\Temporary file"
# folder_path = "."
folder_path = "G:\\Python_Project\\kaoqinghedui\\try"

# print("folder_path = ", folder_path)
xl_sx_files = []
xl_s_files = []
summary_files = []
need_compare_files = []
need_del_files = []
compare_summary_sheet = "核对结果.xlsx"
not_need_files = "备份"

# 存储需要的列数
Compare_indices = {'姓名': None, '全勤': None, '出勤': None, '平时': None, '周末': None, '法定': None, '晚餐补贴': None, '迟到': None,
                   '事假': None, '病假': None, '年假': None}
# 存储系统汇总表需要的列数
system_indices = {'姓名': None, '全勤': None, '实出勤天数': None, '加班1.5': None, '加班2.0': None, '加班3.0': None, '夜班次数': None,
                  '迟到次数': None, '事假天数': None, '病假天数': None, '年休假天数': None, '转调休加班': None}
# 存储汇总表右边的标题
titleList_right = ['姓名', '全勤', '实出勤天数', '平时', '周末', '法定', '晚餐补贴', '迟到', '事假', '病假', '年假']
# 存储汇总表左边的标题
titleList_left = ['姓名', '全勤', '实出勤天数', '加班1.5', '加班2.0', '加班3.0', '夜班次数', '迟到次数', '事假天数', '病假天数', '年休假天数', '转调休加班']
# 存储要复制的信息
data = []
data_name = []

data_len = len(titleList_left)
# 总表实出勤天数
system_workDay = 31

error_print = []


# 异常退出
def quit_print():
    print('-----------------------------------------------------------------------')
    print('-----------------------------注意警告/错误信息--------------------------')
    for log in error_print:
        print(log)

    for l_need_del_file in need_del_files:
        os.remove(os.path.join(folder_path, l_need_del_file))
    sys.exit()


# 获取所需要的行数（根据名字）
def from_name_get_need_row(indices, file_name):
    source_wb = load_workbook(os.path.join(folder_path, file_name), data_only=True)
    source_sheet = source_wb.active

    for row in source_sheet.iter_rows():
        for cell in row:
            for header in indices.keys():
                if str(cell.value) in header:
                    indices[header] = cell.row

    source_wb.close()
    # print(indices)
    return indices


# 获取所需要的行数（根据工号列，找到包含2200的行）
def get_need_row(file_name):

    source_wb = load_workbook(os.path.join(folder_path, file_name), data_only=True)

    sheet_count = len(source_wb.sheetnames)     # 查看有多少个工作表，正常应该是1个
    if sheet_count > 1:
        error_print.append(f"{file_name}有多个工作表")

    source_sheet = source_wb.worksheets[0]
    # 包含工号的列序号
    job_num_cells = []
    # 包含工号的列序号的行号
    job_num_cells_row = []
    # 包含2200的列序号
    job_num_2200_cells = []
    # 真正包含工号和2200的列序号
    job_num_2200_cell = None
    # 真正需要的行数
    need_row = []
    # 记录当前工号含2200的最大次数
    max_2200 = 0
    # 记录当前工号含2200的最大次数的序号数
    max_2200_num = 0

    # 找到所有包含'工号'的列
    for row in source_sheet.iter_rows():
        if not row:
            error_print.append(f"{file_name}行数据有问题")
            return 0
        for cell in row:
            if '工号' in str(cell.value):
                job_num_cells.append(row.index(cell))
                job_num_cells_row.append(cell.row)

    # 在每一个'工号'列中寻找'2200'
    tmp = 0
    same_max = 0
    if job_num_cells is None:
        error_print.append(f"{file_name}没有找到工号")
        return 0
    for i, job_num_cell in enumerate(job_num_cells):
        count = 0
        for row in source_sheet.iter_rows(min_row=job_num_cells_row[i]):
            # print(job_num_cells_row[i])
            cell = row[job_num_cell]
            # print(str(cell.value))
            if re.search(r'2\d00', str(cell.value)):
                count += 1
        if count >= 1:
            if count == max_2200:
                same_max = max_2200
            if count > max_2200:
                max_2200 = count
                max_2200_num = tmp

            job_num_2200_cells.append(job_num_cell)
            tmp += 1

    # 判断job_num_2200_cells的元素个数，返回相应的列序号
    if len(job_num_2200_cells) == 0:
        error_print.append(f"{file_name}工号列下面找不到2200工号列有{job_num_cells}")
        return 0
    elif len(job_num_2200_cells) >= 2:
        if same_max == max_2200:
            error_print.append(f"{file_name}有多个工号列序号切2200数量一样，返回默认第2列，对应表格第3列")
            job_num_2200_cell = 2  # 默认返回C列的序号
        else:
            error_print.append(f"{file_name}有多个工号列序号，返回默认数量最多的")
            job_num_2200_cell = job_num_2200_cells[max_2200_num]  # 默认返回C列的序号
            error_print.append(f"{file_name}max_2200_num{max_2200_num},job_num_2200_cells[]{job_num_2200_cell}")
    elif len(job_num_2200_cells) == 1:
        job_num_2200_cell = job_num_2200_cells[0]  # 返回列表中的元素作为列序号
    else:
        pass

    # 根据真正的工号列找需要的行数
    start_row = None
    for row in source_sheet.iter_rows():
        cell = row[job_num_2200_cell]
        if '工号' in str(cell.value):
            start_row = cell.row

    if not start_row or start_row is None:
        error_print.append(f"{file_name}当前工号列{job_num_2200_cell}找不到'工号'")
        return 0
    for row in source_sheet.iter_rows(min_row=start_row+1):
        cell = row[job_num_2200_cell]
        if re.search(r'2\d00', str(cell.value)):
            need_row.append(cell.row)

    source_wb.close()
    if need_row is None:
        error_print.append(f"{file_name}的工号列{job_num_2200_cell}找不到2200")
        return 0

    return need_row


# 根据获取需要的列数（根据所需要的标题，获取列数，只遍历rows_with_job最前一行之前内容（目的是为了找标题））
def get_need_cell(rows_with_job, indices, file_name):
    source_wb = load_workbook(os.path.join(folder_path, file_name), data_only=True)
    source_sheet = source_wb.worksheets[0]
    del_flag = 0
    for row in source_sheet.iter_rows(max_row=min(rows_with_job)-1, values_only=True):
        for cell in row:
            for header in indices.keys():
                if header in str(cell):
                    indices[header] = row.index(cell) + 1

    for key, value in indices.items():
        if value == "" or value is None or value == [] or value == {}:
            error_print.append(f"{file_name}, Key: {key}, Value: {value}")
            if key == "晚餐补贴":
                indices[key] = 100
            else:
                del_flag = 1

    if del_flag == 1:
        quit_print()

    # indices = {header: 1000 if value is None else value for header, value in indices.items()}
    source_wb.close()
    # print(indices)
    pass
    return 1


# （创建的工作簿， 工作表， 需要核对的文件名）
def compare_fun(workbook, sheet_name_in_compare_fun, file_name):
    # 清空字典并将值设置为None
    for key in Compare_indices:
        Compare_indices[key] = None

    # 获取所需要的行数（根据工号）
    rows_with_job = get_need_row(file_name)

    if rows_with_job is None or rows_with_job == 0:
        return 0

    # 获取需要的列数
    if not get_need_cell(rows_with_job, Compare_indices, file_name):
        return 0

    # # 给表创建一个标题
    ws = workbook[sheet_name_in_compare_fun]
    for i, value in enumerate(titleList_left):
        ws.cell(row=1, column=i + 1, value=value)
    for i, value in enumerate(titleList_right):
        ws.cell(row=1, column=i + 1 + len(titleList_left) + 1, value=value)
    # ws = workbook.create_sheet(title=sheet_name_in_compare_fun)
    # ws.append(titleList_left + titleList_right)

    # 获取单元格内容并赋值到另一个表
    source_wb = load_workbook(os.path.join(folder_path, file_name), data_only=True)
    source_sheet = source_wb.worksheets[0]

    start_row = 2
    # 循环把一行单元格数据，放到data，然后粘贴到另一个表
    # print(Compare_indices)
    for row in rows_with_job:
        for cell in Compare_indices.values():
            # print(row,cell)
            if cell == list(Compare_indices.values())[0]:
                data_name.append(source_sheet.cell(row=row, column=cell).value)

            data.append(source_sheet.cell(row=row, column=cell).value)
        # print(data)
        for i, value in enumerate(data):
            if value is None:
                ws.cell(row=start_row, column=i + 1 + data_len + 1, value=0)
            else:
                ws.cell(row=start_row, column=i + 1 + data_len + 1, value=value)
        start_row += 1
        data.clear()
    source_wb.close()
    pass

    # 根据名字创建字典
    data_name_indices = {item: None for item in data_name}
    # print(data_name_indices)

    # 根据名字获取系统表的行列，返回列表
    rows_with_name_summary_dist = from_name_get_need_row(data_name_indices, summary_files[0])
    # print(rows_with_name_summary_dist)
    rows_with_name_summary_bak = dict(from_name_get_need_row(data_name_indices, summary_files[0]))  # 备份未被删除时要获取的行数
    # print(rows_with_name_summary_bak)

    del_flag = 0
    keys_to_delete = []  # 用于存储需要删除的键
    for key, value in rows_with_name_summary_dist.items():
        if value is None:
            error_print.append(f"{file_name}键 '{key}' 的行数为空，请检查总表是否有这个人")
            keys_to_delete.append(key)
            del_flag = 1

    for key in keys_to_delete:
        del rows_with_name_summary_dist[key]

    if del_flag == 1:
        # quit_print()
        pass

    rows_with_name_summary = rows_with_name_summary_dist.values()
    # if rows_with_name_summary is None:
    #     # 如果 rows_with_name_summary 是 None，执行一些替代的操作
    #     print("rows_with_name_summary 是 None，进行替代处理")
    #     # 这里添加替代处理的代码

    # 根据所需要的标题，获取列数，只遍历rows_with_job（前面获取了包含2200的行）最前一行之前内容（目的是为了找标题），根据标题找列
    get_need_cell(rows_with_name_summary, system_indices, summary_files[0])

    # 打开总表
    source_wb = load_workbook(os.path.join(folder_path, summary_files[0]), data_only=True)
    source_sheet = source_wb.worksheets[0]

    start_row = 2
    # 循环把一行单元格数据，放到data，然后粘贴到另一个表
    # system_workDay = input("实出勤天数：  输入0表示系统表获取  输入其他数字作为输入天数")

    for row in rows_with_name_summary_bak.values():
        if row is None:
            ws.cell(row=start_row, column=1, value="无此人")
            start_row += 1
            continue
        for key, cell in system_indices.items():
            data.append(source_sheet.cell(row=row, column=cell).value)
        # print(data)
        i = 0
        for i, value in enumerate(data):
            ws.cell(row=start_row, column=i + 1, value=value)
        start_row += 1
        data.clear()
    source_wb.close()
    pass
    data_name.clear()
    workbook.save(os.path.join(folder_path, compare_summary_sheet))
    return 1


# 去掉路径和文件名后缀，只保留文件名
def get_filename_without_extension(file_path):
    # 去掉路径
    file_name = os.path.basename(file_path)
    # 去掉后缀
    return os.path.splitext(file_name)[0]


# 获取需要对比的文件名，和总文件名
def get_xls_or_sx_summary_files():
    global need_compare_files
    for file in os.listdir(folder_path):
        if not_need_files in file:
            continue
        if file.endswith(".xls"):
            file = get_filename_without_extension(file)
            # print(file)
            p.save_book_as(file_name=(os.path.join(folder_path, (file + '.xls'))),
                           dest_file_name=(os.path.join(folder_path, (file + '.xlsx'))))
            need_del_files.append(file + '.xlsx')

    for file in os.listdir(folder_path):
        if file == compare_summary_sheet or not_need_files in file:
            continue
        if file.endswith(".xlsx"):
            if "考勤汇总" in file:
                summary_files.append(file)
            else:
                if file.endswith(".xlsx"):
                    xl_sx_files.append(file)
                else:
                    xl_s_files.append(file)
    need_compare_files = xl_sx_files + xl_s_files
    if need_compare_files:
        print("需要核对的文件：", need_compare_files)
    else:
        for get_need_del_file in need_del_files:
            os.remove(os.path.join(folder_path, get_need_del_file))
        error_print.append("没有找到部门文件")
        quit_print()

    if summary_files:
        print("系统导出的总文件：", summary_files)
    else:
        for get_need_del_file in need_del_files:
            os.remove(os.path.join(folder_path, get_need_del_file))
        error_print.append("没有找到系统总表")
        quit_print()


# 创建工作簿
def compare_summary_file_create():
    file_path = os.path.join(folder_path, compare_summary_sheet)    # 文件全路径
    base_name = get_filename_without_extension(compare_summary_sheet)   # 文件名，无前缀后缀
    backup_count = 1    # 备份计数

    # 检查文件是否存在
    if os.path.exists(file_path):
        print(f"文件 {compare_summary_sheet} 已存在，备份！")
        while True:
            # 构造新的备份文件名
            backup_name = f"{base_name}备份_{backup_count}.xlsx"
            new_file_path = os.path.join(folder_path, backup_name)

            # 如果新文件名也存在，递增备份编号并重试
            if os.path.exists(new_file_path):
                backup_count += 1
                continue

            # 加载现有的工作簿，并保存到新的备份文件名
            work = load_workbook(file_path)
            work.save(new_file_path)
            print(f"备份已创建：{backup_name}")
            break

    work = Workbook()
    print("创建工作簿：", compare_summary_sheet)

    return work


# 创建核对工作表（创建的工作簿，需要核对的文件名）
def compare_summary_sheet_create(workbook, file_name):
    # 获取所有工作表名字
    sheet_names = workbook.sheetnames
    # print("工作簿里的表有：", sheet_names)
    # 获取要创建的工作表名字（去掉文件名后缀）
    sheet_name_creat = get_filename_without_extension(file_name) + "核对"
    # 不存在则创建工作表
    if sheet_name_creat in sheet_names:
        # print("工作表已存在：", sheet_name_creat)
        # workbook.close()
        # return 0
        workbook.remove(workbook[sheet_name_creat])

    workbook.create_sheet(sheet_name_creat)
    # print("创建工作表", sheet_name_creat)

    # 删除空表
    for delSheet in sheet_names:
        if 'Sheet' in delSheet:
            workbook.remove(workbook[delSheet])

    print(f"正在复制数据到[{sheet_name_creat}]中。。。")
    return sheet_name_creat


def compare_summary_fun():
    last_wb = load_workbook(os.path.join(folder_path, compare_summary_sheet))

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    black_font = Font(color="000000")
    white_font = Font(color="FFFFFF")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # print("into compare_summary_fun")
    for sheet in last_wb.worksheets:
        for row in sheet.iter_rows(min_row=2):
            # 对比每一对列
            for i in range(1, 11):  # 从A列到K列，总共11列
                if row[i].value != row[i + data_len + 1].value:  # A列和M列的索引差为12，B列和N列的索引差为12，以此类推
                    if i == 2:
                        if row[i].value == row[i + data_len + 1].value+1:
                            continue
                    if row[i].value is None and row[i + data_len + 1].value == 0:
                        continue
                    # 如果不一致，将整行标黄
                    # print(row[i].value, row[i + 12].value)
                    # for cell in row:
                    #     cell.fill = yellow_fill
                    # 将不一致的单元格标红
                    # 标红左边
                    # row[i].fill = red_fill
                    # row[i].font = white_font
                    # 标红右边
                    row[i + data_len + 1].fill = red_fill
                    row[i + data_len + 1].font = white_font

    for sheet in last_wb.worksheets:
        for row in sheet.iter_rows(min_row=2):

            val_3 = row[3].value if row[3].value is not None else 0
            val_4 = row[4].value if row[4].value is not None else 0
            val_5 = row[5].value if row[5].value is not None else 0

            # 获取需要检查的值
            val_11 = row[11].value
            val_16 = row[16].value
            val_17 = row[17].value
            val_18 = row[18].value
            str_flag = 0
            # 检查这些值是否是字符串
            if isinstance(val_11, str):
                error_print.append(f"{sheet} row[11] 是字符串: {val_11}")
                str_flag = 1
            if isinstance(val_16, str):
                error_print.append(f"{sheet} row[16] 是字符串: {val_16}")
                str_flag = 1
            if isinstance(val_17, str):
                error_print.append(f"{sheet} row[17] 是字符串: {val_17}")
                str_flag = 1
            if isinstance(val_18, str):
                error_print.append(f"{sheet} row[18] 是字符串: {val_18}")
                str_flag = 1

            if str_flag == 1:
                quit_print()

            if row[11].value == row[16].value + row[17].value + row[18].value:
                if row[16].value+row[17].value+row[18].value == 0:
                    if val_3 + val_4 + val_5 != 0:
                        continue
                row[3].fill = no_fill
                row[3].font = black_font
                # row[3].border = thin_border

                row[4].fill = no_fill
                row[4].font = black_font
                # row[4].border = thin_border

                row[3 + data_len + 1].fill = no_fill
                row[3 + data_len + 1].font = black_font
                # row[3 + data_len + 1].border = thin_border

                row[4 + data_len + 1].fill = no_fill
                row[4 + data_len + 1].font = black_font
                # row[4 + data_len + 1].border = thin_border

    last_wb.save(os.path.join(folder_path, compare_summary_sheet))


print(folder_path)
get_xls_or_sx_summary_files()       # 检索目录下的xls，xls转xlsx，并添加到要删除的文件列表里，获取所有xlsx文件，备份除外，获取要核对的文件，系统总文件
wb = compare_summary_file_create()      # 创建核对结果文件，如果存在则将其改名为备份1,2，3以此类推。返回核对结果文件
for need_compare_file in need_compare_files:        # 遍历每一个需要核对的部门文件
    sheet_name = compare_summary_sheet_create(wb, need_compare_file)        # 根据部门文件去掉前后缀创建工作表，并删除原来工作簿的sheet空表，返回创建的表
    compare_fun(wb, sheet_name, need_compare_file)         # 当前部门文件与总文件对比

compare_summary_fun()
for need_del_file in need_del_files:
    os.remove(os.path.join(folder_path, need_del_file))

print('-----------------------------------------------------------------------')
print('-----------------------------注意警告/错误信息--------------------------')
for error in error_print:
    print(error)
input("对比完成，按任意键结束。。")
